from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook


class UnsupportedFileError(ValueError):
    pass


def read_file_bytes(field_file) -> bytes:
    field_file.open("rb")
    try:
        return field_file.read()
    finally:
        field_file.close()


def read_excel_rows(field_file) -> list[tuple[str, list[list[object]]]]:
    suffix = Path(field_file.name).suffix.lower()
    data = read_file_bytes(field_file)

    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
        result: list[tuple[str, list[list[object]]]] = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            result.append((sheet.title, rows))
        return result

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise UnsupportedFileError(
                "Reading XLS requires xlrd. Install dependencies from requirements.txt."
            ) from exc

        workbook = xlrd.open_workbook(file_contents=data)
        result = []
        for sheet in workbook.sheets():
            rows = [sheet.row_values(idx) for idx in range(sheet.nrows)]
            result.append((sheet.name, rows))
        return result

    raise UnsupportedFileError("Only XLSX and XLS files are supported.")


def read_pdf_tables(field_file) -> list[tuple[str, list[list[object]]]]:
    data = read_file_bytes(field_file)
    result: list[tuple[str, list[list[object]]]] = []
    with pdfplumber.open(BytesIO(data)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables(
                {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 4,
                    "join_tolerance": 4,
                    "intersection_tolerance": 5,
                }
            )
            for table_number, table in enumerate(tables or [], start=1):
                if table:
                    result.append((f"Page {page_number}, table {table_number}", table))
    return result
