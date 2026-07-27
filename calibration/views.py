from __future__ import annotations

from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from operations_portal.decorators import module_access_required
from operations_portal.models import ModuleCode

from .forms import CalibrationCycleForm, CalibrationVerificationForm, InstrumentForm
from .models import (
    CalibrationAlert,
    CalibrationCycle,
    CalibrationCycleStatus,
    Instrument,
    InstrumentStatus,
)
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .import_forms import CalibrationExcelImportForm
from .models import CalibrationCycle, Instrument
from .services.excel_import import (
    calibration_cycle_instrument_field,
    import_calibration_workbook,
    importable_fields,
    instrument_key_fields,
)


@login_required
@module_access_required(ModuleCode.CALIBRATION)
def dashboard(request):
    today = timezone.localdate()
    due_limit = today + timedelta(days=15)
    instruments = Instrument.objects.filter(is_active=True).select_related(
        "category", "department", "custodian", "monitor"
    )
    context = {
        "total_count": instruments.count(),
        "due_soon_count": instruments.filter(
            next_due_date__gte=today, next_due_date__lte=due_limit
        ).count(),
        "overdue_count": instruments.filter(next_due_date__lt=today).count(),
        "at_lab_count": instruments.filter(status=InstrumentStatus.SENT_TO_LAB).count(),
        "due_instruments": instruments.filter(
            next_due_date__lte=due_limit
        ).order_by("next_due_date")[:20],
        "open_alerts": CalibrationAlert.objects.filter(is_resolved=False).select_related(
            "instrument", "assigned_to"
        )[:20],
    }
    return render(request, "calibration/dashboard.html", context)


@login_required
@module_access_required(ModuleCode.CALIBRATION)
def instrument_list(request):
    qs = Instrument.objects.select_related(
        "category", "department", "custodian", "monitor"
    )
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    if query:
        qs = qs.filter(
            Q(asset_number__icontains=query)
            | Q(description__icontains=query)
            | Q(purchase_serial_number__icontains=query)
            | Q(manufacturer_serial_number__icontains=query)
            | Q(location__icontains=query)
        )
    if status:
        qs = qs.filter(status=status)
    return render(
        request,
        "calibration/instrument_list.html",
        {"instruments": qs, "query": query, "status": status, "status_choices": InstrumentStatus.choices},
    )


@login_required
@module_access_required(ModuleCode.CALIBRATION)
def instrument_detail(request, pk):
    instrument = get_object_or_404(
        Instrument.objects.select_related(
            "category", "department", "custodian", "monitor"
        ),
        pk=pk,
    )
    return render(
        request,
        "calibration/instrument_detail.html",
        {
            "instrument": instrument,
            "cycles": instrument.calibration_cycles.select_related("laboratory", "verified_by"),
            "history": instrument.history_events.select_related("performed_by", "calibration_cycle")[:100],
        },
    )


@login_required
@module_access_required(ModuleCode.CALIBRATION)
@permission_required("calibration.add_instrument", raise_exception=True)
def instrument_create(request):
    form = InstrumentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        instrument = form.save(commit=False)
        instrument.created_by = request.user
        instrument.save()
        messages.success(request, "Instrument added to the calibration master list.")
        return redirect("calibration:instrument_detail", pk=instrument.pk)
    return render(request, "calibration/form.html", {"form": form, "title": "Add instrument or gauge"})


@login_required
@module_access_required(ModuleCode.CALIBRATION)
@permission_required("calibration.change_instrument", raise_exception=True)
def instrument_update(request, pk):
    instrument = get_object_or_404(Instrument, pk=pk)
    form = InstrumentForm(request.POST or None, instance=instrument)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Instrument master record updated.")
        return redirect("calibration:instrument_detail", pk=instrument.pk)
    return render(request, "calibration/form.html", {"form": form, "title": "Update instrument"})


@login_required
@module_access_required(ModuleCode.CALIBRATION)
@permission_required("calibration.add_calibrationcycle", raise_exception=True)
def cycle_create(request, instrument_pk):
    instrument = get_object_or_404(Instrument, pk=instrument_pk)
    form = CalibrationCycleForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        cycle = form.save(commit=False)
        cycle.instrument = instrument
        cycle.created_by = request.user
        cycle.save()
        messages.success(request, f"Calibration cycle {cycle.cycle_number} created.")
        return redirect("calibration:instrument_detail", pk=instrument.pk)
    return render(
        request,
        "calibration/form.html",
        {"form": form, "title": f"New calibration cycle - {instrument.asset_number}"},
    )


@login_required
@module_access_required(ModuleCode.CALIBRATION)
@permission_required("calibration.verify_calibration_cycle", raise_exception=True)
def cycle_verify(request, pk):
    cycle = get_object_or_404(CalibrationCycle.objects.select_related("instrument"), pk=pk)
    form = CalibrationVerificationForm(
        request.POST or None,
        request.FILES or None,
        instance=cycle,
    )
    if request.method == "POST" and form.is_valid():
        form.save(verified_by=request.user)
        messages.success(request, "Calibration certificate verified and history card updated.")
        return redirect("calibration:instrument_detail", pk=cycle.instrument_id)
    return render(
        request,
        "calibration/form.html",
        {"form": form, "title": f"QC verification - {cycle.instrument.asset_number}"},
    )


@login_required
@module_access_required(ModuleCode.CALIBRATION)
@permission_required("calibration.print_calibration_master_list", raise_exception=True)
def master_list(request):
    instruments = Instrument.objects.filter(is_active=True).select_related(
        "category", "department", "custodian", "monitor"
    )
    return render(request, "calibration/master_list.html", {"instruments": instruments})


@login_required
@module_access_required(ModuleCode.CALIBRATION)
def due_list(request):
    today = timezone.localdate()
    due_limit = today + timedelta(days=15)
    instruments = Instrument.objects.filter(
        is_active=True,
        next_due_date__lte=due_limit,
    ).select_related("category", "department", "custodian", "monitor").order_by("next_due_date")
    return render(request, "calibration/due_list.html", {"instruments": instruments, "today": today})


def _format_header(field):
    return str(field.verbose_name).strip().title()


def _prepare_template_sheet(
    sheet,
    fields,
    first_headers=None,
):
    first_headers = first_headers or []

    headers = first_headers + [
        _format_header(field)
        for field in fields
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = sheet.cell(
            row=1,
            column=column_number,
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = header_fill

        cell.comment = Comment(
            f"Excel import field: {header}",
            "Quality Management System",
        )

        sheet.column_dimensions[
            get_column_letter(column_number)
        ].width = max(
            18,
            min(len(str(header)) + 5, 40),
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


@staff_member_required
def download_excel_template(request):
    workbook = Workbook()

    instructions = workbook.active
    instructions.title = "Instructions"

    instructions_rows = [
        ["Calibration Excel Import Instructions"],
        [
            "1",
            "Enter the instrument master data in the Instruments sheet.",
        ],
        [
            "2",
            "Enter previous calibration transactions in the "
            "Calibration History sheet.",
        ],
        [
            "3",
            "Do not change the worksheet names or column headings.",
        ],
        [
            "4",
            "Use DD-MM-YYYY or YYYY-MM-DD for dates.",
        ],
        [
            "5",
            "Related values such as Custodian, Monitor and Department "
            "must already exist in the system.",
        ],
        [
            "6",
            "Blank cells do not overwrite existing information.",
        ],
        [
            "7",
            "Actual PDF calibration certificates must be uploaded "
            "separately after import.",
        ],
        [
            "8",
            "Instrument Lookup in Calibration History must contain "
            "the instrument code, purchase serial number, asset code "
            "or serial number.",
        ],
    ]

    for row in instructions_rows:
        instructions.append(row)

    instructions["A1"].font = Font(
        bold=True,
        size=16,
    )

    instructions.column_dimensions["A"].width = 12
    instructions.column_dimensions["B"].width = 90

    instrument_fields = importable_fields(
        Instrument
    )

    instruments_sheet = workbook.create_sheet(
        "Instruments"
    )

    _prepare_template_sheet(
        instruments_sheet,
        instrument_fields,
    )

    key_fields = set(instrument_key_fields())

    for column_number, field in enumerate(
        instrument_fields,
        start=1,
    ):
        cell = instruments_sheet.cell(
            row=1,
            column=column_number,
        )

        details = [
            f"Internal field: {field.name}",
        ]

        if field.name in key_fields:
            details.append(
                "Used to identify and update existing instruments."
            )

        if field.choices:
            details.append(
                "Allowed values: "
                + ", ".join(
                    str(label)
                    for _, label in field.flatchoices
                )
            )

        if field.is_relation:
            details.append(
                "Enter the related record's username, email, "
                "code or name."
            )

        cell.comment = Comment(
            "\n".join(details),
            "Quality Management System",
        )

    relationship_field = (
        calibration_cycle_instrument_field()
    )

    history_fields = importable_fields(
        CalibrationCycle,
        excluded_names={
            relationship_field.name,
        },
    )

    history_sheet = workbook.create_sheet(
        "Calibration History"
    )

    _prepare_template_sheet(
        history_sheet,
        history_fields,
        first_headers=[
            "Instrument Lookup",
        ],
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="calibration_import_template.xlsx"'

    workbook.save(response)

    return response


@staff_member_required
def import_excel(request):
    result = None

    if request.method == "POST":
        form = CalibrationExcelImportForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            result = import_calibration_workbook(
                uploaded_file=form.cleaned_data[
                    "excel_file"
                ],
                update_existing=form.cleaned_data[
                    "update_existing"
                ],
            )

            total_success = (
                result["instruments_created"]
                + result["instruments_updated"]
                + result["history_created"]
                + result["history_updated"]
            )

            if total_success:
                messages.success(
                    request,
                    (
                        f"Excel import completed. "
                        f"{result['instruments_created']} instruments created, "
                        f"{result['instruments_updated']} instruments updated, "
                        f"{result['history_created']} history records created "
                        f"and {result['history_updated']} history records updated."
                    ),
                )

            if not result["errors"]:
                return redirect(
                    "calibration:dashboard"
                )

    else:
        form = CalibrationExcelImportForm()

    return render(
        request,
        "calibration/import_excel.html",
        {
            "form": form,
            "result": result,
        },
    )