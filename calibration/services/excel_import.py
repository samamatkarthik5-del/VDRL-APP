import re
from datetime import date, datetime
from decimal import Decimal

from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.db.models import Q
from openpyxl import load_workbook

from calibration.models import CalibrationCycle, Instrument


INSTRUMENT_KEY_CANDIDATES = [
    "instrument_code",
    "asset_code",
    "purchase_serial_number",
    "serial_number",
    "equipment_number",
    "identification_number",
]

COMMON_HEADER_ALIASES = {
    "instrument id": "instrument_code",
    "instrument no": "instrument_code",
    "instrument number": "instrument_code",
    "asset no": "asset_code",
    "asset number": "asset_code",
    "purchase serial no": "purchase_serial_number",
    "purchase serial number": "purchase_serial_number",
    "serial no": "serial_number",
    "serial number": "serial_number",
    "instrument name": "name",
    "instrument description": "description",
    "calibration frequency": "calibration_frequency_months",
    "frequency months": "calibration_frequency_months",
    "latest calibration": "latest_calibration_date",
    "latest calibration date": "latest_calibration_date",
    "last calibration date": "latest_calibration_date",
    "next due date": "next_due_date",
    "due date": "next_due_date",
    "certificate no": "certificate_number",
    "certificate number": "certificate_number",
    "calibration result": "result",
    "calibration laboratory": "laboratory",
    "external laboratory": "laboratory",
    "sent to lab": "sent_to_lab_date",
    "sent to lab date": "sent_to_lab_date",
    "received from lab": "received_from_lab_date",
    "received from lab date": "received_from_lab_date",
    "returned to production": "returned_to_production_date",
    "returned to production date": "returned_to_production_date",
}


def normalize_header(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def is_blank(value):
    return value is None or (
        isinstance(value, str) and not value.strip()
    )


def importable_fields(model, excluded_names=None):
    excluded_names = set(excluded_names or [])

    excluded_types = (
        models.AutoField,
        models.BigAutoField,
        models.FileField,
        models.ImageField,
    )

    return [
        field
        for field in model._meta.fields
        if field.editable
        and not field.auto_created
        and not isinstance(field, excluded_types)
        and field.name not in excluded_names
    ]


def build_field_map(model, excluded_names=None):
    fields = importable_fields(
        model,
        excluded_names=excluded_names,
    )

    mapping = {}

    for field in fields:
        possible_headers = {
            field.name,
            field.attname,
            str(field.verbose_name),
            field.name.replace("_", " "),
        }

        for header in possible_headers:
            mapping[normalize_header(header)] = field

    for alias, field_name in COMMON_HEADER_ALIASES.items():
        try:
            field = model._meta.get_field(field_name)
        except Exception:
            continue

        if field.name not in set(excluded_names or []):
            mapping[normalize_header(alias)] = field

    return mapping


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
    ]

    for date_format in formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            ).date()
        except ValueError:
            continue

    raise ValueError(
        f"Invalid date '{value}'. Use DD-MM-YYYY or YYYY-MM-DD."
    )


def parse_datetime(value):
    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(
            value,
            datetime.min.time(),
        )

    text = str(value).strip()

    formats = [
        "%Y-%m-%d %H:%M",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
    ]

    for date_format in formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            )
        except ValueError:
            continue

    raise ValueError(
        f"Invalid date/time '{value}'."
    )


def parse_boolean(value):
    if isinstance(value, bool):
        return value

    text = str(value).strip().lower()

    true_values = {
        "1",
        "true",
        "yes",
        "y",
        "active",
    }

    false_values = {
        "0",
        "false",
        "no",
        "n",
        "inactive",
    }

    if text in true_values:
        return True

    if text in false_values:
        return False

    raise ValueError(
        f"Invalid Yes/No value '{value}'."
    )


def resolve_choice(field, value):
    text = str(value).strip()

    for choice_value, choice_label in field.flatchoices:
        if text.lower() == str(choice_value).lower():
            return choice_value

        if text.lower() == str(choice_label).lower():
            return choice_value

    allowed = ", ".join(
        str(label)
        for _, label in field.flatchoices
    )

    raise ValueError(
        f"Invalid value '{value}' for {field.verbose_name}. "
        f"Allowed values: {allowed}"
    )


def resolve_related_object(field, value):
    related_model = field.remote_field.model

    if isinstance(value, related_model):
        return value

    text = str(value).strip()

    if text.isdigit():
        result = related_model.objects.filter(
            pk=int(text),
        ).first()

        if result:
            return result

    candidate_fields = [
        "username",
        "email",
        "employee_id",
        "code",
        "name",
        "instrument_code",
        "asset_code",
        "serial_number",
    ]

    query = Q()
    usable_fields = []

    for field_name in candidate_fields:
        try:
            related_field = related_model._meta.get_field(
                field_name
            )
        except Exception:
            continue

        if isinstance(
            related_field,
            (
                models.CharField,
                models.EmailField,
                models.TextField,
            ),
        ):
            usable_fields.append(field_name)
            query |= Q(
                **{
                    f"{field_name}__iexact": text,
                }
            )

    if usable_fields:
        results = related_model.objects.filter(query)

        if results.count() == 1:
            return results.first()

        if results.count() > 1:
            raise ValueError(
                f"More than one {related_model._meta.verbose_name} "
                f"matches '{text}'."
            )

    raise ValueError(
        f"{related_model._meta.verbose_name.title()} "
        f"'{text}' was not found."
    )


def convert_value(field, value):
    if field.choices:
        return resolve_choice(field, value)

    if field.is_relation:
        return resolve_related_object(
            field,
            value,
        )

    if isinstance(field, models.DateTimeField):
        return parse_datetime(value)

    if isinstance(field, models.DateField):
        return parse_date(value)

    if isinstance(field, models.BooleanField):
        return parse_boolean(value)

    if isinstance(
        field,
        (
            models.IntegerField,
            models.PositiveIntegerField,
            models.PositiveSmallIntegerField,
            models.SmallIntegerField,
        ),
    ):
        return int(value)

    if isinstance(field, models.DecimalField):
        return Decimal(str(value))

    if isinstance(field, models.FloatField):
        return float(value)

    return str(value).strip()


def row_to_model_data(headers, row, field_map):
    data = {}

    for index, raw_value in enumerate(row):
        if index >= len(headers):
            continue

        if is_blank(raw_value):
            continue

        normalized_header = headers[index]
        field = field_map.get(normalized_header)

        if not field:
            continue

        data[field.name] = convert_value(
            field,
            raw_value,
        )

    return data


def instrument_key_fields():
    available_fields = {
        field.name: field
        for field in Instrument._meta.fields
    }

    result = []

    for field_name in INSTRUMENT_KEY_CANDIDATES:
        if field_name in available_fields:
            result.append(field_name)

    for field in Instrument._meta.fields:
        if (
            field.unique
            and not field.primary_key
            and field.name not in result
        ):
            result.append(field.name)

    return result


def find_existing_instrument(data):
    for field_name in instrument_key_fields():
        value = data.get(field_name)

        if is_blank(value):
            continue

        field = Instrument._meta.get_field(field_name)

        lookup = field_name

        if isinstance(
            field,
            (
                models.CharField,
                models.TextField,
                models.EmailField,
            ),
        ):
            lookup = f"{field_name}__iexact"

        instrument = Instrument.objects.filter(
            **{
                lookup: value,
            }
        ).first()

        if instrument:
            return instrument, field_name

    return None, None


def find_instrument_from_lookup(value):
    if is_blank(value):
        raise ValueError(
            "Instrument Lookup is required."
        )

    text = str(value).strip()

    if text.isdigit():
        instrument = Instrument.objects.filter(
            pk=int(text),
        ).first()

        if instrument:
            return instrument

    for field_name in instrument_key_fields():
        field = Instrument._meta.get_field(field_name)

        lookup = field_name

        if isinstance(
            field,
            (
                models.CharField,
                models.TextField,
                models.EmailField,
            ),
        ):
            lookup = f"{field_name}__iexact"

        instrument = Instrument.objects.filter(
            **{
                lookup: text,
            }
        ).first()

        if instrument:
            return instrument

    raise ValueError(
        f"Instrument '{text}' was not found."
    )


def get_sheet(workbook, requested_name):
    requested = normalize_header(requested_name)

    for sheet_name in workbook.sheetnames:
        if normalize_header(sheet_name) == requested:
            return workbook[sheet_name]

    return None


def import_instruments(sheet, update_existing, result):
    headers = [
        normalize_header(cell.value)
        for cell in sheet[1]
    ]

    field_map = build_field_map(Instrument)

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        if all(is_blank(value) for value in row):
            continue

        try:
            with transaction.atomic():
                data = row_to_model_data(
                    headers,
                    row,
                    field_map,
                )

                if not data:
                    raise ValueError(
                        "No recognised instrument columns were found."
                    )

                existing, matched_field = (
                    find_existing_instrument(data)
                )

                if existing:
                    if not update_existing:
                        raise ValueError(
                            "Instrument already exists and "
                            "'Update existing' is not selected."
                        )

                    for field_name, value in data.items():
                        setattr(
                            existing,
                            field_name,
                            value,
                        )

                    existing.full_clean()
                    existing.save()

                    result["instruments_updated"] += 1

                else:
                    instrument = Instrument(**data)
                    instrument.full_clean()
                    instrument.save()

                    result["instruments_created"] += 1

        except (
            ValueError,
            ValidationError,
            TypeError,
        ) as error:
            result["errors"].append(
                f"Instruments row {row_number}: {error}"
            )


def calibration_cycle_instrument_field():
    for field in CalibrationCycle._meta.fields:
        if (
            field.is_relation
            and field.remote_field
            and field.remote_field.model is Instrument
        ):
            return field

    raise ValueError(
        "CalibrationCycle has no relationship to Instrument."
    )


def import_calibration_history(sheet, result):
    relationship_field = (
        calibration_cycle_instrument_field()
    )

    excluded_fields = {
        relationship_field.name,
    }

    field_map = build_field_map(
        CalibrationCycle,
        excluded_names=excluded_fields,
    )

    headers = [
        normalize_header(cell.value)
        for cell in sheet[1]
    ]

    lookup_columns = {
        normalize_header("Instrument Lookup"),
        normalize_header("Instrument Code"),
        normalize_header("Asset Code"),
        normalize_header("Purchase Serial Number"),
        normalize_header("Serial Number"),
    }

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        if all(is_blank(value) for value in row):
            continue

        try:
            with transaction.atomic():
                lookup_value = None

                for index, header in enumerate(headers):
                    if header in lookup_columns:
                        if index < len(row) and not is_blank(
                            row[index]
                        ):
                            lookup_value = row[index]
                            break

                instrument = find_instrument_from_lookup(
                    lookup_value
                )

                data = row_to_model_data(
                    headers,
                    row,
                    field_map,
                )

                data[relationship_field.name] = instrument

                cycle_lookup = {
                    relationship_field.name: instrument,
                }

                for candidate in [
                    "calibration_date",
                    "certificate_number",
                    "sent_to_lab_date",
                ]:
                    if candidate in data:
                        cycle_lookup[candidate] = data[candidate]

                if len(cycle_lookup) > 1:
                    defaults = {
                        key: value
                        for key, value in data.items()
                        if key not in cycle_lookup
                    }

                    cycle, created = (
                        CalibrationCycle.objects.update_or_create(
                            **cycle_lookup,
                            defaults=defaults,
                        )
                    )

                    cycle.full_clean()
                    cycle.save()

                    if created:
                        result["history_created"] += 1
                    else:
                        result["history_updated"] += 1

                else:
                    cycle = CalibrationCycle(**data)
                    cycle.full_clean()
                    cycle.save()

                    result["history_created"] += 1

        except (
            ValueError,
            ValidationError,
            TypeError,
        ) as error:
            result["errors"].append(
                f"Calibration History row {row_number}: {error}"
            )


def import_calibration_workbook(
    uploaded_file,
    update_existing=True,
):
    workbook = load_workbook(
        uploaded_file,
        data_only=True,
    )

    result = {
        "instruments_created": 0,
        "instruments_updated": 0,
        "history_created": 0,
        "history_updated": 0,
        "errors": [],
    }

    instruments_sheet = get_sheet(
        workbook,
        "Instruments",
    )

    history_sheet = get_sheet(
        workbook,
        "Calibration History",
    )

    if instruments_sheet is None:
        result["errors"].append(
            "The workbook must contain a sheet named 'Instruments'."
        )
        return result

    import_instruments(
        instruments_sheet,
        update_existing,
        result,
    )

    if history_sheet is not None:
        import_calibration_history(
            history_sheet,
            result,
        )

    return result
