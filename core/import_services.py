import re

from datetime import (
    date,
    datetime,
    time,
)

from django.contrib.auth import (
    get_user_model,
)

from django.core.exceptions import (
    ValidationError,
)

from django.db import transaction

from django.utils import timezone

from openpyxl import (
    Workbook,
    load_workbook,
)

from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)

from openpyxl.utils import (
    get_column_letter,
)


from .models import (
    CRSComment,
    CRSRegister,
    CustomerVDRLTemplate,
    CustomerVDRLTemplateItem,
    Department,
    DocumentMaster,
    SalesOrderVDRL,
    SalesOrderVDRLDocument,
)


User = get_user_model()


# =========================================================
# IMPORT TYPES
# =========================================================

IMPORT_CUSTOMER_TEMPLATE = (
    "CUSTOMER_TEMPLATE"
)

IMPORT_SALES_ORDER_VDRL = (
    "SALES_ORDER_VDRL"
)

IMPORT_CRS_COMMENTS = (
    "CRS_COMMENTS"
)


# =========================================================
# IMPORT MODES
# =========================================================

MODE_CREATE_ONLY = "CREATE_ONLY"

MODE_UPDATE_EXISTING = (
    "UPDATE_EXISTING"
)

MODE_UPSERT = "UPSERT"


# =========================================================
# HEADERS
# =========================================================

CUSTOMER_TEMPLATE_HEADERS = [
    "SEQUENCE NUMBER",
    "INTERNAL DOCUMENT CODE",
    "CUSTOMER DOCUMENT TITLE",
    "REQUIREMENT TYPE",
    "CONDITION DESCRIPTION",
    "SUBMISSION STAGE",
    "DUE DATE BASIS",
    "DAY OFFSET",
    "CUSTOMER REVIEW DAYS",
    "RESPONSIBLE DEPARTMENT CODE",
    "APPROVAL REQUIRED",
    "CRS REQUIRED",
    "INCLUDE IN FINAL MRB",
    "REQUIRED FILE FORMAT",
    "REMARKS",
    "IS ACTIVE",
]


SALES_ORDER_VDRL_HEADERS = [
    "SEQUENCE NUMBER",
    "INTERNAL DOCUMENT CODE",
    "CUSTOMER DOCUMENT CODE",
    "DOCUMENT TITLE",
    "REQUIREMENT TYPE",
    "CONDITION DESCRIPTION",
    "APPLICABILITY STATUS",
    "SUBMISSION STAGE",
    "DUE DATE BASIS",
    "DAY OFFSET",
    "PLANNED SUBMISSION DATE",
    "FORECAST SUBMISSION DATE",
    "CUSTOMER REVIEW DAYS",
    "RESPONSIBLE DEPARTMENT CODE",
    "RESPONSIBLE USERNAME",
    "CURRENT REVISION",
    "APPROVAL REQUIRED",
    "CRS REQUIRED",
    "INCLUDE IN FINAL MRB",
    "REQUIRED FILE FORMAT",
    "REMARKS",
    "IS ACTIVE",
]


CRS_COMMENT_HEADERS = [
    "COMMENT NUMBER",
    "PAGE REFERENCE",
    "CLAUSE REFERENCE",
    "CUSTOMER COMMENT",
    "CATEGORY",
    "ASSIGNED DEPARTMENT CODE",
    "ASSIGNED USERNAME",
    "DECISION",
    "SUPPLIER RESPONSE",
    "INTERNAL ACTION REQUIRED",
    "DOCUMENT UPDATE STATUS",
    "STATUS",
    "CUSTOMER DISPOSITION",
    "ASSIGNED AT",
    "TARGET RESPONSE DATE",
    "RESPONSE COMPLETED AT",
    "REMARKS",
]


# =========================================================
# BASIC HELPERS
# =========================================================

def normalize_header(value):
    """
    Normalize Excel header text.

    Example:
    ' Internal  Document_Code '
        ->
    'INTERNAL DOCUMENT CODE'
    """

    if value is None:
        return ""

    text = str(
        value
    ).strip()

    text = re.sub(
        r"[_\-]+",
        " ",
        text,
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.upper()


def clean_text(
    value,
    default="",
):
    if value is None:
        return default

    if isinstance(
        value,
        float,
    ):
        if value.is_integer():
            return str(
                int(
                    value
                )
            )

    return str(
        value
    ).strip()


def parse_integer(
    value,
    field_name,
    row_number,
    default=None,
    minimum=None,
):
    if (
        value is None
        or clean_text(value) == ""
    ):
        if default is not None:
            return default

        raise ValueError(
            (
                f"{field_name} is required."
            )
        )

    try:
        number = int(
            float(
                value
            )
        )

    except (
        TypeError,
        ValueError,
    ):
        raise ValueError(
            (
                f"{field_name} must be "
                "a whole number."
            )
        )

    if (
        minimum is not None
        and number < minimum
    ):
        raise ValueError(
            (
                f"{field_name} must be "
                f"{minimum} or greater."
            )
        )

    return number


def parse_optional_integer(
    value,
    field_name,
    minimum=None,
):
    if (
        value is None
        or clean_text(value) == ""
    ):
        return None

    return parse_integer(
        value,
        field_name,
        row_number=None,
        minimum=minimum,
    )


def parse_boolean(
    value,
    default=False,
):
    if (
        value is None
        or clean_text(value) == ""
    ):
        return default

    normalized = (
        clean_text(
            value
        )
        .strip()
        .upper()
    )

    true_values = {
        "TRUE",
        "YES",
        "Y",
        "1",
        "ACTIVE",
    }

    false_values = {
        "FALSE",
        "NO",
        "N",
        "0",
        "INACTIVE",
    }

    if normalized in true_values:
        return True

    if normalized in false_values:
        return False

    raise ValueError(
        (
            f"Invalid Yes/No value: "
            f"{value}"
        )
    )


def choice_lookup(
    choices,
):
    lookup = {}

    for (
        value,
        label,
    ) in choices:

        lookup[
            normalize_header(
                value
            )
        ] = value

        lookup[
            normalize_header(
                label
            )
        ] = value

    return lookup


def parse_choice(
    value,
    choices,
    field_name,
    default=None,
):
    if (
        value is None
        or clean_text(value) == ""
    ):
        if default is not None:
            return default

        raise ValueError(
            (
                f"{field_name} is required."
            )
        )

    lookup = choice_lookup(
        choices
    )

    normalized = normalize_header(
        value
    )

    if normalized not in lookup:
        allowed_values = ", ".join(
            str(
                choice_value
            )
            for (
                choice_value,
                _,
            ) in choices
        )

        raise ValueError(
            (
                f"Invalid {field_name}: "
                f"'{value}'. "
                f"Allowed values: "
                f"{allowed_values}"
            )
        )

    return lookup[
        normalized
    ]


def parse_date_value(
    value,
    field_name,
    required=False,
):
    if (
        value is None
        or clean_text(value) == ""
    ):
        if required:
            raise ValueError(
                (
                    f"{field_name} "
                    "is required."
                )
            )

        return None

    if isinstance(
        value,
        datetime,
    ):
        return value.date()

    if isinstance(
        value,
        date,
    ):
        return value

    text = clean_text(
        value
    )

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d-%b-%Y",
        "%d %b %Y",
    ]

    for date_format in formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            ).date()

        except ValueError:
            continue

    raise ValueError(
        (
            f"Invalid {field_name}: "
            f"'{value}'. "
            "Use YYYY-MM-DD, "
            "DD-MM-YYYY or DD/MM/YYYY."
        )
    )


def parse_datetime_value(
    value,
    field_name,
    required=False,
    default=None,
):
    if (
        value is None
        or clean_text(value) == ""
    ):
        if default is not None:
            return default

        if required:
            raise ValueError(
                (
                    f"{field_name} "
                    "is required."
                )
            )

        return None

    if isinstance(
        value,
        datetime,
    ):
        parsed_value = value

    elif isinstance(
        value,
        date,
    ):
        parsed_value = (
            datetime.combine(
                value,
                time.min,
            )
        )

    else:
        text = clean_text(
            value
        )

        formats = [
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%d-%m-%Y %H:%M",
            "%d-%m-%Y",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
        ]

        parsed_value = None

        for date_format in formats:
            try:
                parsed_value = (
                    datetime.strptime(
                        text,
                        date_format,
                    )
                )

                break

            except ValueError:
                continue

        if parsed_value is None:
            raise ValueError(
                (
                    f"Invalid {field_name}: "
                    f"'{value}'."
                )
            )

    if timezone.is_naive(
        parsed_value
    ):
        parsed_value = (
            timezone.make_aware(
                parsed_value,
                timezone.get_current_timezone(),
            )
        )

    return parsed_value


def flatten_validation_error(
    error,
):
    if hasattr(
        error,
        "message_dict",
    ):
        messages = []

        for (
            field_name,
            field_errors,
        ) in (
            error
            .message_dict
            .items()
        ):
            messages.append(
                (
                    f"{field_name}: "
                    f"{', '.join(field_errors)}"
                )
            )

        return "; ".join(
            messages
        )

    if hasattr(
        error,
        "messages",
    ):
        return "; ".join(
            error.messages
        )

    return str(
        error
    )


# =========================================================
# WORKBOOK READING
# =========================================================

def read_import_rows(
    uploaded_file,
    required_headers,
):
    """
    Read the IMPORT worksheet.

    Row 1 must contain the headers.
    """

    uploaded_file.seek(
        0
    )

    workbook = load_workbook(
        uploaded_file,
        read_only=True,
        data_only=True,
    )

    try:
        if (
            "IMPORT"
            in workbook.sheetnames
        ):
            worksheet = workbook[
                "IMPORT"
            ]

        else:
            worksheet = (
                workbook.active
            )

        raw_headers = [
            cell.value
            for cell
            in next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                )
            )
        ]

        normalized_headers = [
            normalize_header(
                header
            )
            for header
            in raw_headers
        ]

        duplicate_headers = {
            header
            for header
            in normalized_headers
            if (
                header
                and
                normalized_headers.count(
                    header
                ) > 1
            )
        }

        if duplicate_headers:
            raise ValueError(
                (
                    "Duplicate Excel headers: "
                    + ", ".join(
                        sorted(
                            duplicate_headers
                        )
                    )
                )
            )

        missing_headers = [
            header
            for header
            in required_headers
            if (
                header
                not in normalized_headers
            )
        ]

        if missing_headers:
            raise ValueError(
                (
                    "Missing required Excel "
                    "headers: "
                    + ", ".join(
                        missing_headers
                    )
                )
            )

        header_indexes = {
            header: index
            for (
                index,
                header,
            ) in enumerate(
                normalized_headers
            )
            if header
        }

        rows = []

        for (
            row_number,
            values,
        ) in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):

            if all(
                value is None
                or clean_text(
                    value
                ) == ""
                for value
                in values
            ):
                continue

            row_data = {
                header: (
                    values[index]
                    if index < len(
                        values
                    )
                    else None
                )
                for (
                    header,
                    index,
                ) in (
                    header_indexes
                    .items()
                )
            }

            row_data[
                "_ROW_NUMBER"
            ] = row_number

            rows.append(
                row_data
            )

            if len(
                rows
            ) > 5000:
                raise ValueError(
                    (
                        "Maximum import size is "
                        "5,000 data rows."
                    )
                )

        return rows

    finally:
        workbook.close()


# =========================================================
# REFERENCE LOOKUPS
# =========================================================

def get_document_map():
    return {
        (
            document
            .internal_document_code
            .strip()
            .casefold()
        ): document

        for document in (
            DocumentMaster
            .objects
            .filter(
                is_active=True
            )
        )
    }


def get_department_map():
    return {
        (
            department
            .code
            .strip()
            .casefold()
        ): department

        for department in (
            Department
            .objects
            .filter(
                is_active=True
            )
        )
    }


def get_user_map():
    user_map = {}

    for user in (
        User.objects
        .filter(
            is_active=True
        )
    ):
        user_map[
            user
            .username
            .strip()
            .casefold()
        ] = user

        if user.email:
            user_map[
                user
                .email
                .strip()
                .casefold()
            ] = user

    return user_map


def lookup_document(
    value,
    document_map,
):
    code = (
        clean_text(
            value
        )
        .casefold()
    )

    if not code:
        raise ValueError(
            (
                "INTERNAL DOCUMENT CODE "
                "is required."
            )
        )

    document = (
        document_map.get(
            code
        )
    )

    if not document:
        raise ValueError(
            (
                "Unknown INTERNAL DOCUMENT "
                f"CODE: '{value}'."
            )
        )

    return document


def lookup_department(
    value,
    department_map,
    required=False,
):
    code = (
        clean_text(
            value
        )
        .casefold()
    )

    if not code:
        if required:
            raise ValueError(
                (
                    "RESPONSIBLE DEPARTMENT "
                    "CODE is required."
                )
            )

        return None

    department = (
        department_map.get(
            code
        )
    )

    if not department:
        raise ValueError(
            (
                "Unknown department code: "
                f"'{value}'."
            )
        )

    return department


def lookup_user(
    value,
    user_map,
    required=False,
):
    key = (
        clean_text(
            value
        )
        .casefold()
    )

    if not key:
        if required:
            raise ValueError(
                (
                    "Username is required."
                )
            )

        return None

    user = user_map.get(
        key
    )

    if not user:
        raise ValueError(
            (
                "Unknown username or email: "
                f"'{value}'."
            )
        )

    return user


# =========================================================
# RESULT HELPERS
# =========================================================

def empty_result(
    dry_run,
):
    return {
        "success": False,
        "dry_run": dry_run,
        "created": 0,
        "updated": 0,
        "processed": 0,
        "errors": [],
    }


def add_row_error(
    result,
    row_number,
    error_message,
):
    result[
        "errors"
    ].append(
        {
            "row": row_number,
            "message": (
                error_message
            ),
        }
    )


def validate_import_mode(
    existing,
    import_mode,
):
    if (
        import_mode
        == MODE_CREATE_ONLY
        and existing
    ):
        raise ValueError(
            (
                "A matching record already "
                "exists."
            )
        )

    if (
        import_mode
        == MODE_UPDATE_EXISTING
        and not existing
    ):
        raise ValueError(
            (
                "No existing matching "
                "record was found."
            )
        )


# =========================================================
# CUSTOMER TEMPLATE IMPORT
# =========================================================

def import_customer_template_items(
    uploaded_file,
    template,
    import_mode,
    dry_run=False,
):
    result = empty_result(
        dry_run
    )

    try:
        rows = read_import_rows(
            uploaded_file,
            [
                "SEQUENCE NUMBER",
                "INTERNAL DOCUMENT CODE",
            ],
        )

    except Exception as exc:
        add_row_error(
            result,
            "-",
            str(
                exc
            ),
        )

        return result

    document_map = (
        get_document_map()
    )

    department_map = (
        get_department_map()
    )

    prepared_objects = []

    seen_sequences = set()

    for row in rows:
        row_number = row[
            "_ROW_NUMBER"
        ]

        try:
            sequence_number = (
                parse_integer(
                    row.get(
                        "SEQUENCE NUMBER"
                    ),
                    "SEQUENCE NUMBER",
                    row_number,
                    minimum=1,
                )
            )

            if (
                sequence_number
                in seen_sequences
            ):
                raise ValueError(
                    (
                        "Duplicate SEQUENCE "
                        "NUMBER inside the "
                        "uploaded workbook."
                    )
                )

            seen_sequences.add(
                sequence_number
            )

            document = (
                lookup_document(
                    row.get(
                        "INTERNAL DOCUMENT CODE"
                    ),
                    document_map,
                )
            )

            existing = (
                CustomerVDRLTemplateItem
                .objects
                .filter(
                    template=template,
                    sequence_number=(
                        sequence_number
                    ),
                )
                .first()
            )

            validate_import_mode(
                existing,
                import_mode,
            )

            responsible_department = (
                lookup_department(
                    row.get(
                        (
                            "RESPONSIBLE "
                            "DEPARTMENT CODE"
                        )
                    ),
                    department_map,
                )
                or
                document
                .default_responsible_department
            )

            customer_review_days = (
                parse_optional_integer(
                    row.get(
                        "CUSTOMER REVIEW DAYS"
                    ),
                    (
                        "CUSTOMER REVIEW "
                        "DAYS"
                    ),
                    minimum=0,
                )
            )

            data = {
                "template": template,

                "sequence_number": (
                    sequence_number
                ),

                "document": document,

                "customer_document_title": (
                    clean_text(
                        row.get(
                            (
                                "CUSTOMER "
                                "DOCUMENT TITLE"
                            )
                        )
                    )
                ),

                "requirement_type": (
                    parse_choice(
                        row.get(
                            "REQUIREMENT TYPE"
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .RequirementType
                            .choices
                        ),
                        "REQUIREMENT TYPE",
                        default=(
                            CustomerVDRLTemplateItem
                            .RequirementType
                            .MANDATORY
                        ),
                    )
                ),

                "condition_description": (
                    clean_text(
                        row.get(
                            (
                                "CONDITION "
                                "DESCRIPTION"
                            )
                        )
                    )
                ),

                "submission_stage": (
                    parse_choice(
                        row.get(
                            "SUBMISSION STAGE"
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .SubmissionStage
                            .choices
                        ),
                        "SUBMISSION STAGE",
                        default=(
                            CustomerVDRLTemplateItem
                            .SubmissionStage
                            .AFTER_ORDER
                        ),
                    )
                ),

                "due_date_basis": (
                    parse_choice(
                        row.get(
                            "DUE DATE BASIS"
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .DueDateBasis
                            .choices
                        ),
                        "DUE DATE BASIS",
                        default=(
                            CustomerVDRLTemplateItem
                            .DueDateBasis
                            .ORDER_DATE
                        ),
                    )
                ),

                "day_offset": (
                    parse_integer(
                        row.get(
                            "DAY OFFSET"
                        ),
                        "DAY OFFSET",
                        row_number,
                        default=0,
                    )
                ),

                "customer_review_days": (
                    customer_review_days
                ),

                "responsible_department": (
                    responsible_department
                ),

                "approval_required": (
                    parse_boolean(
                        row.get(
                            "APPROVAL REQUIRED"
                        ),
                        default=True,
                    )
                ),

                "crs_required": (
                    parse_boolean(
                        row.get(
                            "CRS REQUIRED"
                        ),
                        default=True,
                    )
                ),

                "include_in_final_mrb": (
                    parse_boolean(
                        row.get(
                            (
                                "INCLUDE IN "
                                "FINAL MRB"
                            )
                        ),
                        default=True,
                    )
                ),

                "required_file_format": (
                    parse_choice(
                        row.get(
                            (
                                "REQUIRED "
                                "FILE FORMAT"
                            )
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .FileFormat
                            .choices
                        ),
                        (
                            "REQUIRED "
                            "FILE FORMAT"
                        ),
                        default=(
                            CustomerVDRLTemplateItem
                            .FileFormat
                            .PDF
                        ),
                    )
                ),

                "remarks": clean_text(
                    row.get(
                        "REMARKS"
                    )
                ),

                "is_active": (
                    parse_boolean(
                        row.get(
                            "IS ACTIVE"
                        ),
                        default=True,
                    )
                ),
            }

            if existing:
                import_object = (
                    existing
                )

                action = "UPDATE"

            else:
                import_object = (
                    CustomerVDRLTemplateItem()
                )

                action = "CREATE"

            for (
                field_name,
                field_value,
            ) in data.items():
                setattr(
                    import_object,
                    field_name,
                    field_value,
                )

            import_object.full_clean()

            prepared_objects.append(
                (
                    action,
                    import_object,
                )
            )

        except (
            ValueError,
            ValidationError,
        ) as exc:
            add_row_error(
                result,
                row_number,
                flatten_validation_error(
                    exc
                ),
            )

    result[
        "processed"
    ] = len(
        rows
    )

    if result["errors"]:
        return result

    result["created"] = sum(
        1
        for (
            action,
            _,
        ) in prepared_objects
        if action == "CREATE"
    )

    result["updated"] = sum(
        1
        for (
            action,
            _,
        ) in prepared_objects
        if action == "UPDATE"
    )

    if dry_run:
        result["success"] = True

        return result

    with transaction.atomic():
        for (
            _,
            import_object,
        ) in prepared_objects:
            import_object.save()

    result["success"] = True

    return result


# =========================================================
# SALES ORDER VDRL IMPORT
# =========================================================

def import_sales_order_vdrl_documents(
    uploaded_file,
    vdrl,
    import_mode,
    dry_run=False,
):
    result = empty_result(
        dry_run
    )

    try:
        rows = read_import_rows(
            uploaded_file,
            [
                "SEQUENCE NUMBER",
                "INTERNAL DOCUMENT CODE",
            ],
        )

    except Exception as exc:
        add_row_error(
            result,
            "-",
            str(
                exc
            ),
        )

        return result

    document_map = (
        get_document_map()
    )

    department_map = (
        get_department_map()
    )

    user_map = (
        get_user_map()
    )

    prepared_objects = []

    seen_sequences = set()

    for row in rows:
        row_number = row[
            "_ROW_NUMBER"
        ]

        try:
            sequence_number = (
                parse_integer(
                    row.get(
                        "SEQUENCE NUMBER"
                    ),
                    "SEQUENCE NUMBER",
                    row_number,
                    minimum=1,
                )
            )

            if (
                sequence_number
                in seen_sequences
            ):
                raise ValueError(
                    (
                        "Duplicate SEQUENCE "
                        "NUMBER inside the "
                        "uploaded workbook."
                    )
                )

            seen_sequences.add(
                sequence_number
            )

            document = (
                lookup_document(
                    row.get(
                        "INTERNAL DOCUMENT CODE"
                    ),
                    document_map,
                )
            )

            existing = (
                SalesOrderVDRLDocument
                .objects
                .filter(
                    vdrl=vdrl,
                    sequence_number=(
                        sequence_number
                    ),
                )
                .first()
            )

            validate_import_mode(
                existing,
                import_mode,
            )

            responsible_department = (
                lookup_department(
                    row.get(
                        (
                            "RESPONSIBLE "
                            "DEPARTMENT CODE"
                        )
                    ),
                    department_map,
                )
                or
                document
                .default_responsible_department
            )

            responsible_person = (
                lookup_user(
                    row.get(
                        (
                            "RESPONSIBLE "
                            "USERNAME"
                        )
                    ),
                    user_map,
                )
            )

            customer_review_days = (
                parse_optional_integer(
                    row.get(
                        "CUSTOMER REVIEW DAYS"
                    ),
                    (
                        "CUSTOMER REVIEW "
                        "DAYS"
                    ),
                    minimum=0,
                )
            )

            if (
                customer_review_days
                is None
            ):
                customer_review_days = (
                    vdrl
                    .sales_order
                    .customer
                    .standard_review_days
                )

            current_revision = (
                clean_text(
                    row.get(
                        "CURRENT REVISION"
                    )
                )
            )

            if not current_revision:
                if existing:
                    current_revision = (
                        existing
                        .current_revision
                    )

                else:
                    current_revision = "0"

            data = {
                "vdrl": vdrl,

                "sequence_number": (
                    sequence_number
                ),

                "document": document,

                "customer_document_code": (
                    clean_text(
                        row.get(
                            (
                                "CUSTOMER "
                                "DOCUMENT CODE"
                            )
                        )
                    )
                ),

                "document_title": (
                    clean_text(
                        row.get(
                            "DOCUMENT TITLE"
                        )
                    )
                    or
                    document.document_title
                ),

                "requirement_type": (
                    parse_choice(
                        row.get(
                            "REQUIREMENT TYPE"
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .RequirementType
                            .choices
                        ),
                        "REQUIREMENT TYPE",
                        default=(
                            CustomerVDRLTemplateItem
                            .RequirementType
                            .MANDATORY
                        ),
                    )
                ),

                "condition_description": (
                    clean_text(
                        row.get(
                            (
                                "CONDITION "
                                "DESCRIPTION"
                            )
                        )
                    )
                ),

                "applicability_status": (
                    parse_choice(
                        row.get(
                            (
                                "APPLICABILITY "
                                "STATUS"
                            )
                        ),
                        (
                            SalesOrderVDRLDocument
                            .ApplicabilityStatus
                            .choices
                        ),
                        (
                            "APPLICABILITY "
                            "STATUS"
                        ),
                        default=(
                            SalesOrderVDRLDocument
                            .ApplicabilityStatus
                            .REQUIRED
                        ),
                    )
                ),

                "submission_stage": (
                    parse_choice(
                        row.get(
                            "SUBMISSION STAGE"
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .SubmissionStage
                            .choices
                        ),
                        "SUBMISSION STAGE",
                        default=(
                            CustomerVDRLTemplateItem
                            .SubmissionStage
                            .AFTER_ORDER
                        ),
                    )
                ),

                "due_date_basis": (
                    parse_choice(
                        row.get(
                            "DUE DATE BASIS"
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .DueDateBasis
                            .choices
                        ),
                        "DUE DATE BASIS",
                        default=(
                            CustomerVDRLTemplateItem
                            .DueDateBasis
                            .ORDER_DATE
                        ),
                    )
                ),

                "day_offset": (
                    parse_integer(
                        row.get(
                            "DAY OFFSET"
                        ),
                        "DAY OFFSET",
                        row_number,
                        default=0,
                    )
                ),

                "planned_submission_date": (
                    parse_date_value(
                        row.get(
                            (
                                "PLANNED "
                                "SUBMISSION DATE"
                            )
                        ),
                        (
                            "PLANNED "
                            "SUBMISSION DATE"
                        ),
                    )
                ),

                "forecast_submission_date": (
                    parse_date_value(
                        row.get(
                            (
                                "FORECAST "
                                "SUBMISSION DATE"
                            )
                        ),
                        (
                            "FORECAST "
                            "SUBMISSION DATE"
                        ),
                    )
                ),

                "customer_review_days": (
                    customer_review_days
                ),

                "responsible_department": (
                    responsible_department
                ),

                "responsible_person": (
                    responsible_person
                ),

                "current_revision": (
                    current_revision
                ),

                "approval_required": (
                    parse_boolean(
                        row.get(
                            "APPROVAL REQUIRED"
                        ),
                        default=True,
                    )
                ),

                "crs_required": (
                    parse_boolean(
                        row.get(
                            "CRS REQUIRED"
                        ),
                        default=True,
                    )
                ),

                "include_in_final_mrb": (
                    parse_boolean(
                        row.get(
                            (
                                "INCLUDE IN "
                                "FINAL MRB"
                            )
                        ),
                        default=True,
                    )
                ),

                "required_file_format": (
                    parse_choice(
                        row.get(
                            (
                                "REQUIRED "
                                "FILE FORMAT"
                            )
                        ),
                        (
                            CustomerVDRLTemplateItem
                            .FileFormat
                            .choices
                        ),
                        (
                            "REQUIRED "
                            "FILE FORMAT"
                        ),
                        default=(
                            CustomerVDRLTemplateItem
                            .FileFormat
                            .PDF
                        ),
                    )
                ),

                "remarks": (
                    clean_text(
                        row.get(
                            "REMARKS"
                        )
                    )
                ),

                "is_active": (
                    parse_boolean(
                        row.get(
                            "IS ACTIVE"
                        ),
                        default=True,
                    )
                ),
            }

            if existing:
                import_object = existing

                action = "UPDATE"

            else:
                import_object = (
                    SalesOrderVDRLDocument()
                )

                import_object.source_template_item = (
                    None
                )

                action = "CREATE"

            for (
                field_name,
                field_value,
            ) in data.items():
                setattr(
                    import_object,
                    field_name,
                    field_value,
                )

            if (
                import_object
                .planned_submission_date
                is None
            ):
                (
                    import_object
                    .planned_submission_date
                ) = (
                    import_object
                    .calculate_planned_submission_date()
                )

            import_object.full_clean()

            prepared_objects.append(
                (
                    action,
                    import_object,
                )
            )

        except (
            ValueError,
            ValidationError,
        ) as exc:
            add_row_error(
                result,
                row_number,
                flatten_validation_error(
                    exc
                ),
            )

    result[
        "processed"
    ] = len(
        rows
    )

    if result["errors"]:
        return result

    result["created"] = sum(
        1
        for (
            action,
            _,
        ) in prepared_objects
        if action == "CREATE"
    )

    result["updated"] = sum(
        1
        for (
            action,
            _,
        ) in prepared_objects
        if action == "UPDATE"
    )

    if dry_run:
        result["success"] = True

        return result

    with transaction.atomic():
        for (
            _,
            import_object,
        ) in prepared_objects:
            import_object.save()

    result["success"] = True

    return result


# =========================================================
# CRS COMMENT IMPORT
# =========================================================

def import_crs_comments(
    uploaded_file,
    crs,
    import_mode,
    imported_by,
    dry_run=False,
):
    result = empty_result(
        dry_run
    )

    try:
        rows = read_import_rows(
            uploaded_file,
            [
                "COMMENT NUMBER",
                "CUSTOMER COMMENT",
            ],
        )

    except Exception as exc:
        add_row_error(
            result,
            "-",
            str(
                exc
            ),
        )

        return result

    department_map = (
        get_department_map()
    )

    user_map = (
        get_user_map()
    )

    prepared_objects = []

    seen_comment_numbers = set()

    for row in rows:
        row_number = row[
            "_ROW_NUMBER"
        ]

        try:
            comment_number = (
                clean_text(
                    row.get(
                        "COMMENT NUMBER"
                    )
                )
            )

            if not comment_number:
                raise ValueError(
                    (
                        "COMMENT NUMBER "
                        "is required."
                    )
                )

            comment_key = (
                comment_number
                .casefold()
            )

            if (
                comment_key
                in seen_comment_numbers
            ):
                raise ValueError(
                    (
                        "Duplicate COMMENT NUMBER "
                        "inside the uploaded "
                        "workbook."
                    )
                )

            seen_comment_numbers.add(
                comment_key
            )

            customer_comment = (
                clean_text(
                    row.get(
                        "CUSTOMER COMMENT"
                    )
                )
            )

            if not customer_comment:
                raise ValueError(
                    (
                        "CUSTOMER COMMENT "
                        "is required."
                    )
                )

            existing = (
                CRSComment
                .objects
                .filter(
                    crs=crs,
                    comment_number=(
                        comment_number
                    ),
                )
                .first()
            )

            validate_import_mode(
                existing,
                import_mode,
            )

            assigned_department = (
                lookup_department(
                    row.get(
                        (
                            "ASSIGNED "
                            "DEPARTMENT CODE"
                        )
                    ),
                    department_map,
                )
            )

            assigned_person = (
                lookup_user(
                    row.get(
                        (
                            "ASSIGNED "
                            "USERNAME"
                        )
                    ),
                    user_map,
                )
            )

            assigned_at = (
                parse_datetime_value(
                    row.get(
                        "ASSIGNED AT"
                    ),
                    "ASSIGNED AT",
                    default=(
                        existing.assigned_at
                        if existing
                        else timezone.now()
                    ),
                )
            )

            data = {
                "crs": crs,

                "comment_number": (
                    comment_number
                ),

                "page_reference": (
                    clean_text(
                        row.get(
                            "PAGE REFERENCE"
                        )
                    )
                ),

                "clause_reference": (
                    clean_text(
                        row.get(
                            "CLAUSE REFERENCE"
                        )
                    )
                ),

                "customer_comment": (
                    customer_comment
                ),

                "category": (
                    parse_choice(
                        row.get(
                            "CATEGORY"
                        ),
                        (
                            CRSComment
                            .Category
                            .choices
                        ),
                        "CATEGORY",
                        default=(
                            CRSComment
                            .Category
                            .TECHNICAL
                        ),
                    )
                ),

                "assigned_department": (
                    assigned_department
                ),

                "assigned_person": (
                    assigned_person
                ),

                "decision": (
                    parse_choice(
                        row.get(
                            "DECISION"
                        ),
                        (
                            CRSComment
                            .Decision
                            .choices
                        ),
                        "DECISION",
                        default=(
                            CRSComment
                            .Decision
                            .PENDING
                        ),
                    )
                ),

                "supplier_response": (
                    clean_text(
                        row.get(
                            "SUPPLIER RESPONSE"
                        )
                    )
                ),

                "internal_action_required": (
                    clean_text(
                        row.get(
                            (
                                "INTERNAL "
                                "ACTION REQUIRED"
                            )
                        )
                    )
                ),

                "document_update_status": (
                    parse_choice(
                        row.get(
                            (
                                "DOCUMENT "
                                "UPDATE STATUS"
                            )
                        ),
                        (
                            CRSComment
                            .DocumentUpdateStatus
                            .choices
                        ),
                        (
                            "DOCUMENT "
                            "UPDATE STATUS"
                        ),
                        default=(
                            CRSComment
                            .DocumentUpdateStatus
                            .PENDING
                        ),
                    )
                ),

                "status": (
                    parse_choice(
                        row.get(
                            "STATUS"
                        ),
                        (
                            CRSComment
                            .Status
                            .choices
                        ),
                        "STATUS",
                        default=(
                            CRSComment
                            .Status
                            .OPEN
                        ),
                    )
                ),

                "customer_disposition": (
                    parse_choice(
                        row.get(
                            (
                                "CUSTOMER "
                                "DISPOSITION"
                            )
                        ),
                        (
                            CRSComment
                            .CustomerDisposition
                            .choices
                        ),
                        (
                            "CUSTOMER "
                            "DISPOSITION"
                        ),
                        default=(
                            CRSComment
                            .CustomerDisposition
                            .PENDING
                        ),
                    )
                ),

                "assigned_at": (
                    assigned_at
                ),

                "target_response_date": (
                    parse_date_value(
                        row.get(
                            (
                                "TARGET "
                                "RESPONSE DATE"
                            )
                        ),
                        (
                            "TARGET "
                            "RESPONSE DATE"
                        ),
                    )
                ),

                "response_completed_at": (
                    parse_datetime_value(
                        row.get(
                            (
                                "RESPONSE "
                                "COMPLETED AT"
                            )
                        ),
                        (
                            "RESPONSE "
                            "COMPLETED AT"
                        ),
                    )
                ),

                "remarks": (
                    clean_text(
                        row.get(
                            "REMARKS"
                        )
                    )
                ),

                "updated_by": (
                    imported_by
                ),
            }

            if existing:
                import_object = (
                    existing
                )

                action = "UPDATE"

            else:
                import_object = (
                    CRSComment(
                        created_by=(
                            imported_by
                        )
                    )
                )

                action = "CREATE"

            for (
                field_name,
                field_value,
            ) in data.items():
                setattr(
                    import_object,
                    field_name,
                    field_value,
                )

            import_object.full_clean()

            prepared_objects.append(
                (
                    action,
                    import_object,
                )
            )

        except (
            ValueError,
            ValidationError,
        ) as exc:
            add_row_error(
                result,
                row_number,
                flatten_validation_error(
                    exc
                ),
            )

    result[
        "processed"
    ] = len(
        rows
    )

    if result["errors"]:
        return result

    result["created"] = sum(
        1
        for (
            action,
            _,
        ) in prepared_objects
        if action == "CREATE"
    )

    result["updated"] = sum(
        1
        for (
            action,
            _,
        ) in prepared_objects
        if action == "UPDATE"
    )

    if dry_run:
        result["success"] = True

        return result

    with transaction.atomic():
        for (
            _,
            import_object,
        ) in prepared_objects:
            import_object.save()

        if (
            prepared_objects
            and
            crs.status
            == CRSRegister.Status.DRAFT
        ):
            crs.status = (
                CRSRegister
                .Status
                .IN_PROGRESS
            )

            crs.save(
                update_fields=[
                    "status",
                    "updated_at",
                ]
            )

    result["success"] = True

    return result


# =========================================================
# DOWNLOADABLE EXCEL TEMPLATES
# =========================================================

def style_import_sheet(
    worksheet,
):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[
        1
    ]:
        cell.fill = (
            header_fill
        )

        cell.font = (
            header_font
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = (
        "A2"
    )

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for (
        column_number,
        cell,
    ) in enumerate(
        worksheet[1],
        start=1,
    ):
        width = max(
            len(
                str(
                    cell.value
                )
            ) + 4,
            18,
        )

        worksheet.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = min(
            width,
            35,
        )


def write_choice_reference(
    worksheet,
    title,
    choices,
    start_row,
):
    worksheet.cell(
        row=start_row,
        column=1,
        value=title,
    ).font = Font(
        bold=True
    )

    current_row = (
        start_row + 1
    )

    for (
        value,
        label,
    ) in choices:
        worksheet.cell(
            row=current_row,
            column=1,
            value=value,
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value=label,
        )

        current_row += 1

    return (
        current_row + 1
    )


def build_import_template(
    import_type,
):
    workbook = Workbook()

    import_sheet = (
        workbook.active
    )

    import_sheet.title = (
        "IMPORT"
    )

    if (
        import_type
        == IMPORT_CUSTOMER_TEMPLATE
    ):
        headers = (
            CUSTOMER_TEMPLATE_HEADERS
        )

        instructions = [
            (
                "One row represents one "
                "Customer VDRL Template item."
            ),
            (
                "SEQUENCE NUMBER and "
                "INTERNAL DOCUMENT CODE "
                "are mandatory."
            ),
            (
                "CUSTOMER DOCUMENT CODE "
                "is intentionally not part "
                "of this template."
            ),
        ]

    elif (
        import_type
        == IMPORT_SALES_ORDER_VDRL
    ):
        headers = (
            SALES_ORDER_VDRL_HEADERS
        )

        instructions = [
            (
                "One row represents one "
                "actual Sales Order VDRL "
                "document."
            ),
            (
                "SEQUENCE NUMBER and "
                "INTERNAL DOCUMENT CODE "
                "are mandatory."
            ),
            (
                "Workflow status and "
                "transaction history are "
                "not imported."
            ),
        ]

    elif (
        import_type
        == IMPORT_CRS_COMMENTS
    ):
        headers = (
            CRS_COMMENT_HEADERS
        )

        instructions = [
            (
                "One row represents one "
                "customer CRS comment."
            ),
            (
                "COMMENT NUMBER and "
                "CUSTOMER COMMENT "
                "are mandatory."
            ),
        ]

    else:
        raise ValueError(
            "Unknown import template type."
        )

    import_sheet.append(
        headers
    )

    style_import_sheet(
        import_sheet
    )

    instruction_sheet = (
        workbook.create_sheet(
            "INSTRUCTIONS"
        )
    )

    instruction_sheet.append(
        [
            "IMPORT INSTRUCTIONS"
        ]
    )

    instruction_sheet[
        "A1"
    ].font = Font(
        bold=True
    )

    for instruction in instructions:
        instruction_sheet.append(
            [
                instruction
            ]
        )

    instruction_sheet.append(
        []
    )

    instruction_sheet.append(
        [
            (
                "Boolean values accepted: "
                "YES / NO, TRUE / FALSE, "
                "1 / 0"
            )
        ]
    )

    instruction_sheet.append(
        [
            (
                "Recommended date format: "
                "YYYY-MM-DD"
            )
        ]
    )

    instruction_sheet.append(
        [
            (
                "Do not rename the IMPORT "
                "sheet or header names."
            )
        ]
    )

    instruction_sheet.column_dimensions[
        "A"
    ].width = 100

    reference_sheet = (
        workbook.create_sheet(
            "REFERENCE"
        )
    )

    current_row = 1

    current_row = (
        write_choice_reference(
            reference_sheet,
            "REQUIREMENT TYPE",
            (
                CustomerVDRLTemplateItem
                .RequirementType
                .choices
            ),
            current_row,
        )
    )

    current_row = (
        write_choice_reference(
            reference_sheet,
            "SUBMISSION STAGE",
            (
                CustomerVDRLTemplateItem
                .SubmissionStage
                .choices
            ),
            current_row,
        )
    )

    current_row = (
        write_choice_reference(
            reference_sheet,
            "DUE DATE BASIS",
            (
                CustomerVDRLTemplateItem
                .DueDateBasis
                .choices
            ),
            current_row,
        )
    )

    current_row = (
        write_choice_reference(
            reference_sheet,
            "FILE FORMAT",
            (
                CustomerVDRLTemplateItem
                .FileFormat
                .choices
            ),
            current_row,
        )
    )

    if (
        import_type
        == IMPORT_SALES_ORDER_VDRL
    ):
        current_row = (
            write_choice_reference(
                reference_sheet,
                "APPLICABILITY STATUS",
                (
                    SalesOrderVDRLDocument
                    .ApplicabilityStatus
                    .choices
                ),
                current_row,
            )
        )

    if (
        import_type
        == IMPORT_CRS_COMMENTS
    ):
        for (
            title,
            choices,
        ) in [
            (
                "CATEGORY",
                CRSComment.Category.choices,
            ),
            (
                "DECISION",
                CRSComment.Decision.choices,
            ),
            (
                (
                    "DOCUMENT "
                    "UPDATE STATUS"
                ),
                (
                    CRSComment
                    .DocumentUpdateStatus
                    .choices
                ),
            ),
            (
                "STATUS",
                CRSComment.Status.choices,
            ),
            (
                (
                    "CUSTOMER "
                    "DISPOSITION"
                ),
                (
                    CRSComment
                    .CustomerDisposition
                    .choices
                ),
            ),
        ]:
            current_row = (
                write_choice_reference(
                    reference_sheet,
                    title,
                    choices,
                    current_row,
                )
            )

    reference_sheet.column_dimensions[
        "A"
    ].width = 35

    reference_sheet.column_dimensions[
        "B"
    ].width = 45

    return workbook