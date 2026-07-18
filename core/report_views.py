from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import (
    login_required,
    permission_required,
)
from django.db.models import Avg, Count, Max, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    AuditLog,
    CRSComment,
    Customer,
    Department,
    DocumentTransaction,
    SalesOrder,
    SalesOrderVDRLDocument,
)

from .access import (
    filter_documents_for_user,
)

from .audit import (
    record_audit_event,
)

User = get_user_model()


CUSTOMER_STATUSES = [
    SalesOrderVDRLDocument.DocumentStatus.SUBMITTED,
    SalesOrderVDRLDocument.DocumentStatus.UNDER_CUSTOMER_REVIEW,
    SalesOrderVDRLDocument.DocumentStatus.RESUBMITTED,
]


COMPLETED_STATUSES = [
    SalesOrderVDRLDocument.DocumentStatus.APPROVED,
    SalesOrderVDRLDocument.DocumentStatus.NOT_APPLICABLE,
    SalesOrderVDRLDocument.DocumentStatus.CANCELLED,
]


def get_document_overdue_query(today):
    """
    A document is overdue when:

    1. The required first submission date has passed and
       the document has not yet been submitted.

    OR

    2. The document is with the customer and the customer
       review due date has passed.
    """

    return (
        Q(
            applicability_status=(
                SalesOrderVDRLDocument
                .ApplicabilityStatus
                .REQUIRED
            ),
            first_submission_at__isnull=True,
            planned_submission_date__lt=today,
        )
        |
        Q(
            status__in=CUSTOMER_STATUSES,
            customer_review_due_date__lt=today,
        )
    )


def get_report_filters(request):
    """
    Read report filters from the URL query string.
    """

    return {
        "customer": request.GET.get(
            "customer",
            "",
        ).strip(),

        "sales_order": request.GET.get(
            "sales_order",
            "",
        ).strip(),

        "department": request.GET.get(
            "department",
            "",
        ).strip(),

        "responsible_person": request.GET.get(
            "responsible_person",
            "",
        ).strip(),

        "status": request.GET.get(
            "status",
            "",
        ).strip(),
    }


def apply_document_filters(
    queryset,
    filters,
):
    """
    Apply common filters to a VDRL document queryset.
    """

    if filters["customer"]:
        queryset = queryset.filter(
            vdrl__sales_order__customer_id=(
                filters["customer"]
            )
        )

    if filters["sales_order"]:
        queryset = queryset.filter(
            vdrl__sales_order_id=(
                filters["sales_order"]
            )
        )

    if filters["department"]:
        queryset = queryset.filter(
            responsible_department_id=(
                filters["department"]
            )
        )

    if filters["responsible_person"]:
        queryset = queryset.filter(
            responsible_person_id=(
                filters["responsible_person"]
            )
        )

    if filters["status"]:
        queryset = queryset.filter(
            status=filters["status"]
        )

    return queryset


def get_base_documents():
    """
    Return active documents belonging to the current VDRL
    of active Sales Orders.
    """

    return (
        SalesOrderVDRLDocument.objects
        .filter(
            is_active=True,
            vdrl__is_current=True,
            vdrl__sales_order__is_active=True,
        )
        .select_related(
            "vdrl",
            "vdrl__sales_order",
            "vdrl__sales_order__customer",
            "vdrl__sales_order__project",
            "document",
            "responsible_department",
            "responsible_person",
        )
    )


def build_employee_performance(
    documents,
    today,
):
    """
    Build employee performance using completed internal elapsed periods,
    current pending documents and CRS comments.
    """

    transaction_rows = (
        DocumentTransaction.objects
        .filter(
            document__in=documents,
            elapsed_holder_type=(
                DocumentTransaction
                .HolderType
                .INTERNAL
            ),
            elapsed_responsible_person__isnull=False,
        )
        .values(
            "elapsed_responsible_person_id"
        )
        .annotate(
            completed_periods=Count("id"),

            total_internal_days=Sum(
                "elapsed_calendar_days"
            ),

            average_internal_days=Avg(
                "elapsed_calendar_days"
            ),

            maximum_internal_days=Max(
                "elapsed_calendar_days"
            ),
        )
    )

    transaction_data = {
        row[
            "elapsed_responsible_person_id"
        ]: row

        for row in transaction_rows
    }


    internal_pending_documents = (
        documents
        .exclude(
            status__in=(
                CUSTOMER_STATUSES
                + COMPLETED_STATUSES
                + [
                    SalesOrderVDRLDocument
                    .DocumentStatus
                    .ON_HOLD
                ]
            )
        )
        .filter(
            responsible_person__isnull=False
        )
        .values(
            "responsible_person_id"
        )
        .annotate(
            current_pending=Count("id")
        )
    )

    pending_data = {
        row["responsible_person_id"]:
        row["current_pending"]

        for row in internal_pending_documents
    }


    overdue_internal_documents = (
        documents
        .filter(
            responsible_person__isnull=False,
            first_submission_at__isnull=True,
            planned_submission_date__lt=today,
        )
        .values(
            "responsible_person_id"
        )
        .annotate(
            overdue_internal=Count("id")
        )
    )

    overdue_internal_data = {
        row["responsible_person_id"]:
        row["overdue_internal"]

        for row in overdue_internal_documents
    }


    open_crs_comments = (
        CRSComment.objects
        .filter(
            crs__document__in=documents,
            assigned_person__isnull=False,
        )
        .exclude(
            status=CRSComment.Status.CLOSED
        )
    )


    crs_rows = (
        open_crs_comments
        .values(
            "assigned_person_id"
        )
        .annotate(
            open_crs_comments=Count("id"),

            overdue_crs_comments=Count(
                "id",
                filter=Q(
                    target_response_date__lt=today
                ),
            ),
        )
    )

    crs_data = {
        row["assigned_person_id"]: row
        for row in crs_rows
    }


    employee_ids = set(
        transaction_data.keys()
    )

    employee_ids.update(
        pending_data.keys()
    )

    employee_ids.update(
        overdue_internal_data.keys()
    )

    employee_ids.update(
        crs_data.keys()
    )


    users = (
        User.objects
        .filter(
            id__in=employee_ids
        )
        .select_related(
            "employee_profile",
            "employee_profile__department",
        )
        .order_by(
            "first_name",
            "last_name",
            "username",
        )
    )


    employee_rows = []

    for user in users:
        transaction_row = (
            transaction_data.get(
                user.id,
                {},
            )
        )

        crs_row = (
            crs_data.get(
                user.id,
                {},
            )
        )

        try:
            employee_profile = (
                user.employee_profile
            )

        except Exception:
            employee_profile = None


        if (
            employee_profile
            and employee_profile.department
        ):
            department_name = (
                employee_profile
                .department
                .name
            )

        else:
            department_name = "-"


        employee_rows.append(
            {
                "user": user,

                "department_name": (
                    department_name
                ),

                "completed_periods": (
                    transaction_row.get(
                        "completed_periods",
                        0,
                    )
                    or 0
                ),

                "total_internal_days": (
                    transaction_row.get(
                        "total_internal_days",
                        Decimal("0.00"),
                    )
                    or Decimal("0.00")
                ),

                "average_internal_days": (
                    transaction_row.get(
                        "average_internal_days",
                        Decimal("0.00"),
                    )
                    or Decimal("0.00")
                ),

                "maximum_internal_days": (
                    transaction_row.get(
                        "maximum_internal_days",
                        Decimal("0.00"),
                    )
                    or Decimal("0.00")
                ),

                "current_pending": (
                    pending_data.get(
                        user.id,
                        0,
                    )
                ),

                "overdue_internal": (
                    overdue_internal_data.get(
                        user.id,
                        0,
                    )
                ),

                "open_crs_comments": (
                    crs_row.get(
                        "open_crs_comments",
                        0,
                    )
                    or 0
                ),

                "overdue_crs_comments": (
                    crs_row.get(
                        "overdue_crs_comments",
                        0,
                    )
                    or 0
                ),
            }
        )

    return employee_rows


def build_report_data(request):
    """
    Build all datasets used by both the HTML report page
    and the Excel export.
    """

    today = timezone.localdate()

    filters = get_report_filters(
        request
    )

    documents = apply_document_filters(
        get_base_documents(),
        filters,
    )

    documents = (
        filter_documents_for_user(
            request.user,
            documents,
        )
    )


    overdue_documents = (
        documents
        .filter(
            get_document_overdue_query(
                today
            )
        )
        .distinct()
        .order_by(
            "planned_submission_date",
            "customer_review_due_date",
            "sequence_number",
        )
    )


    customer_pending_documents = (
        documents
        .filter(
            status__in=CUSTOMER_STATUSES
        )
        .order_by(
            "customer_review_due_date",
            "sequence_number",
        )
    )


    internal_pending_documents = (
        documents
        .exclude(
            status__in=(
                CUSTOMER_STATUSES
                + COMPLETED_STATUSES
                + [
                    SalesOrderVDRLDocument
                    .DocumentStatus
                    .ON_HOLD
                ]
            )
        )
        .order_by(
            "planned_submission_date",
            "sequence_number",
        )
    )


    open_crs_comments = (
        CRSComment.objects
        .filter(
            crs__document__in=documents
        )
        .exclude(
            status=CRSComment.Status.CLOSED
        )
        .select_related(
            "crs",
            "crs__document",
            "crs__document__vdrl",
            "crs__document__vdrl__sales_order",
            "crs__document__vdrl__sales_order__customer",
            "assigned_department",
            "assigned_person",
        )
        .order_by(
            "target_response_date",
            "assigned_at",
        )
    )


    employee_rows = (
        build_employee_performance(
            documents,
            today,
        )
    )


    sales_order_summary_queryset = (
        documents
        .values(
            "vdrl__sales_order_id",
            "vdrl__sales_order__sales_order_number",
            "vdrl__sales_order__customer__name",
            "vdrl__sales_order__project__project_name",
        )
        .annotate(
            total_documents=Count("id"),

            approved_documents=Count(
                "id",
                filter=Q(
                    status=(
                        SalesOrderVDRLDocument
                        .DocumentStatus
                        .APPROVED
                    )
                ),
            ),

            customer_pending=Count(
                "id",
                filter=Q(
                    status__in=(
                        CUSTOMER_STATUSES
                    )
                ),
            ),

            internal_pending=Count(
                "id",
                filter=(
                    ~Q(
                        status__in=(
                            CUSTOMER_STATUSES
                            + COMPLETED_STATUSES
                            + [
                                SalesOrderVDRLDocument
                                .DocumentStatus
                                .ON_HOLD
                            ]
                        )
                    )
                ),
            ),
        )
        .order_by(
            "vdrl__sales_order__sales_order_number"
        )
    )


    sales_order_summary = []

    for row in sales_order_summary_queryset:
        total_documents = (
            row["total_documents"]
            or 0
        )

        approved_documents = (
            row["approved_documents"]
            or 0
        )

        if total_documents:
            progress_percent = round(
                approved_documents
                * 100
                / total_documents
            )

        else:
            progress_percent = 0


        row["progress_percent"] = (
            progress_percent
        )

        sales_order_summary.append(
            row
        )


    summary = {
        "total_documents": (
            documents.count()
        ),

        "approved_documents": (
            documents
            .filter(
                status=(
                    SalesOrderVDRLDocument
                    .DocumentStatus
                    .APPROVED
                )
            )
            .count()
        ),

        "customer_pending": (
            customer_pending_documents
            .count()
        ),

        "internal_pending": (
            internal_pending_documents
            .count()
        ),

        "overdue_documents": (
            overdue_documents
            .count()
        ),

        "open_crs_comments": (
            open_crs_comments
            .count()
        ),

        "overdue_crs_comments": (
            open_crs_comments
            .filter(
                target_response_date__lt=today
            )
            .count()
        ),
    }


    return {
        "today": today,
        "filters": filters,
        "documents": documents,
        "overdue_documents": overdue_documents,
        "customer_pending_documents": (
            customer_pending_documents
        ),
        "internal_pending_documents": (
            internal_pending_documents
        ),
        "open_crs_comments": open_crs_comments,
        "employee_rows": employee_rows,
        "sales_order_summary": (
            sales_order_summary
        ),
        "summary": summary,
    }


@login_required
@permission_required(
    "core.view_management_reports",
    raise_exception=True,
)
def management_reports(request):
    """
    Main management reporting screen.
    """

    report_data = build_report_data(
        request
    )

    context = {
        **report_data,

        "customers": (
            Customer.objects
            .filter(is_active=True)
            .order_by("name")
        ),

        "sales_orders": (
            SalesOrder.objects
            .filter(is_active=True)
            .select_related(
                "customer"
            )
            .order_by(
                "-order_date",
                "sales_order_number",
            )
        ),

        "departments": (
            Department.objects
            .filter(is_active=True)
            .order_by("name")
        ),

        "users": (
            User.objects
            .filter(is_active=True)
            .order_by(
                "first_name",
                "last_name",
                "username",
            )
        ),

        "status_choices": (
            SalesOrderVDRLDocument
            .DocumentStatus
            .choices
        ),

        "filter_query": (
            request.GET.urlencode()
        ),
    }

    return render(
        request,
        "core/reports.html",
        context,
    )


def format_excel_datetime(value):
    """
    Convert date/time values into readable Excel-safe text.
    """

    if not value:
        return ""

    if getattr(value, "tzinfo", None) is not None:
        value = timezone.localtime(
            value
        )

    if hasattr(
        value,
        "strftime",
    ):
        if hasattr(
            value,
            "hour",
        ):
            return value.strftime(
                "%d-%b-%Y %H:%M"
            )

        return value.strftime(
            "%d-%b-%Y"
        )

    return str(value)


def user_display_name(user):
    if not user:
        return ""

    return (
        user.get_full_name().strip()
        or user.username
    )


def autosize_worksheet(
    worksheet,
    maximum_width=50,
):
    """
    Automatically adjust Excel column widths.
    """

    for column_cells in (
        worksheet.columns
    ):
        maximum_length = 0

        column_letter = (
            get_column_letter(
                column_cells[0].column
            )
        )

        for cell in column_cells:
            try:
                cell_length = len(
                    str(
                        cell.value
                        if cell.value is not None
                        else ""
                    )
                )

                maximum_length = max(
                    maximum_length,
                    cell_length,
                )

            except (
                TypeError,
                ValueError,
            ):
                continue

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            maximum_length + 2,
            maximum_width,
        )


def write_report_sheet(
    workbook,
    title,
    headers,
    rows,
):
    """
    Create and style one report worksheet.
    """

    worksheet = workbook.create_sheet(
        title=title[:31]
    )

    worksheet.append(
        headers
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )


    for cell in worksheet[1]:
        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


    for row in rows:
        worksheet.append(
            row
        )


    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    autosize_worksheet(
        worksheet
    )

    return worksheet


def get_document_excel_row(
    document,
):
    return [
        (
            document
            .vdrl
            .sales_order
            .sales_order_number
        ),

        (
            document
            .vdrl
            .sales_order
            .customer
            .name
        ),

        (
            document
            .vdrl
            .sales_order
            .project
            .project_name

            if (
                document
                .vdrl
                .sales_order
                .project
            )

            else ""
        ),

        document.sequence_number,

        document.customer_document_code,

        (
            document
            .document
            .internal_document_code
        ),

        document.document_title,

        (
            document
            .get_applicability_status_display()
        ),

        format_excel_datetime(
            document
            .planned_submission_date
        ),

        format_excel_datetime(
            document
            .forecast_submission_date
        ),

        document.current_revision,

        document.current_cycle,

        document.get_status_display(),

        document.current_holder,

        (
            document
            .responsible_department
            .name

            if (
                document
                .responsible_department
            )

            else ""
        ),

        user_display_name(
            document.responsible_person
        ),

        format_excel_datetime(
            document.first_submission_at
        ),

        format_excel_datetime(
            document.last_submission_at
        ),

        format_excel_datetime(
            document
            .last_customer_return_at
        ),

        format_excel_datetime(
            document
            .customer_review_due_date
        ),

        format_excel_datetime(
            document.final_approval_at
        ),

        float(
            document
            .current_aging_days
        ),

        float(
            document
            .total_internal_days
        ),

        float(
            document
            .total_customer_days
        ),

        document.remarks,
    ]


@login_required
@permission_required(
    "core.view_management_reports",
    raise_exception=True,
)
def export_management_reports_xlsx(
    request,
):
    """
    Export the filtered management report to one Excel workbook.
    """

    report_data = build_report_data(
        request
    )


    workbook = Workbook()

    summary_sheet = workbook.active

    summary_sheet.title = "Summary"


    summary_rows = [
        [
            "Generated At",
            timezone.localtime(
                timezone.now()
            ).strftime(
                "%d-%b-%Y %H:%M"
            ),
        ],

        [
            "Total Documents",
            report_data[
                "summary"
            ][
                "total_documents"
            ],
        ],

        [
            "Approved Documents",
            report_data[
                "summary"
            ][
                "approved_documents"
            ],
        ],

        [
            "Customer Pending",
            report_data[
                "summary"
            ][
                "customer_pending"
            ],
        ],

        [
            "Internal Pending",
            report_data[
                "summary"
            ][
                "internal_pending"
            ],
        ],

        [
            "Overdue Documents",
            report_data[
                "summary"
            ][
                "overdue_documents"
            ],
        ],

        [
            "Open CRS Comments",
            report_data[
                "summary"
            ][
                "open_crs_comments"
            ],
        ],

        [
            "Overdue CRS Comments",
            report_data[
                "summary"
            ][
                "overdue_crs_comments"
            ],
        ],
    ]


    summary_sheet.append(
        [
            "KPI",
            "Value",
        ]
    )


    for row in summary_rows:
        summary_sheet.append(
            row
        )


    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )


    for cell in summary_sheet[1]:
        cell.fill = header_fill

        cell.font = header_font


    autosize_worksheet(
        summary_sheet
    )


    document_headers = [
        "Sales Order",
        "Customer",
        "Project",
        "Sequence",
        "Customer Document Code",
        "Internal Document Code",
        "Document Title",
        "Applicability",
        "Planned Submission Date",
        "Forecast Submission Date",
        "Current Revision",
        "Cycle",
        "Status",
        "Current Holder",
        "Responsible Department",
        "Responsible Person",
        "First Submission",
        "Latest Submission",
        "Latest Customer Return",
        "Customer Review Due",
        "Final Approval",
        "Current Aging Days",
        "Total Internal Days",
        "Total Customer Days",
        "Remarks",
    ]


    write_report_sheet(
        workbook,
        "VDRL Status",
        document_headers,
        [
            get_document_excel_row(
                document
            )
            for document in (
                report_data[
                    "documents"
                ]
                .order_by(
                    "vdrl__sales_order__sales_order_number",
                    "sequence_number",
                )
            )
        ],
    )


    write_report_sheet(
        workbook,
        "Overdue Documents",
        document_headers,
        [
            get_document_excel_row(
                document
            )
            for document in (
                report_data[
                    "overdue_documents"
                ]
            )
        ],
    )


    write_report_sheet(
        workbook,
        "Customer Pending",
        document_headers,
        [
            get_document_excel_row(
                document
            )
            for document in (
                report_data[
                    "customer_pending_documents"
                ]
            )
        ],
    )


    write_report_sheet(
        workbook,
        "Internal Pending",
        document_headers,
        [
            get_document_excel_row(
                document
            )
            for document in (
                report_data[
                    "internal_pending_documents"
                ]
            )
        ],
    )


    employee_headers = [
        "Employee",
        "Department",
        "Completed Internal Periods",
        "Total Internal Days",
        "Average Internal Days",
        "Maximum Internal Days",
        "Current Pending Documents",
        "Overdue Internal Documents",
        "Open CRS Comments",
        "Overdue CRS Comments",
    ]


    employee_excel_rows = []

    for row in report_data[
        "employee_rows"
    ]:
        employee_excel_rows.append(
            [
                user_display_name(
                    row["user"]
                ),

                row[
                    "department_name"
                ],

                row[
                    "completed_periods"
                ],

                float(
                    row[
                        "total_internal_days"
                    ]
                ),

                round(
                    float(
                        row[
                            "average_internal_days"
                        ]
                    ),
                    2,
                ),

                round(
                    float(
                        row[
                            "maximum_internal_days"
                        ]
                    ),
                    2,
                ),

                row[
                    "current_pending"
                ],

                row[
                    "overdue_internal"
                ],

                row[
                    "open_crs_comments"
                ],

                row[
                    "overdue_crs_comments"
                ],
            ]
        )


    write_report_sheet(
        workbook,
        "Employee Performance",
        employee_headers,
        employee_excel_rows,
    )


    crs_headers = [
        "Sales Order",
        "Customer",
        "Document",
        "CRS Reference",
        "Cycle",
        "Comment Number",
        "Customer Comment",
        "Assigned Department",
        "Assigned Person",
        "Decision",
        "Supplier Response",
        "Status",
        "Target Response Date",
        "Aging Days",
        "Overdue",
        "Customer Disposition",
    ]


    crs_excel_rows = []

    for comment in report_data[
        "open_crs_comments"
    ]:
        crs_excel_rows.append(
            [
                (
                    comment
                    .crs
                    .document
                    .vdrl
                    .sales_order
                    .sales_order_number
                ),

                (
                    comment
                    .crs
                    .document
                    .vdrl
                    .sales_order
                    .customer
                    .name
                ),

                (
                    comment
                    .crs
                    .document
                    .document_title
                ),

                comment.crs.crs_reference,

                comment.crs.cycle_number,

                comment.comment_number,

                comment.customer_comment,

                (
                    comment
                    .assigned_department
                    .name

                    if (
                        comment
                        .assigned_department
                    )

                    else ""
                ),

                user_display_name(
                    comment.assigned_person
                ),

                comment.get_decision_display(),

                comment.supplier_response,

                comment.get_status_display(),

                format_excel_datetime(
                    comment
                    .target_response_date
                ),

                float(
                    comment.aging_days
                ),

                (
                    "Yes"
                    if comment.is_overdue
                    else "No"
                ),

                (
                    comment
                    .get_customer_disposition_display()
                ),
            ]
        )


    write_report_sheet(
        workbook,
        "CRS Aging",
        crs_headers,
        crs_excel_rows,
    )


    sales_order_headers = [
        "Sales Order",
        "Customer",
        "Project",
        "Total Documents",
        "Approved",
        "With Customer",
        "Internal Pending",
        "Progress %",
    ]


    sales_order_excel_rows = []

    for row in report_data[
        "sales_order_summary"
    ]:
        sales_order_excel_rows.append(
            [
                row[
                    "vdrl__sales_order__sales_order_number"
                ],

                row[
                    "vdrl__sales_order__customer__name"
                ],

                (
                    row[
                        "vdrl__sales_order__project__project_name"
                    ]
                    or ""
                ),

                row[
                    "total_documents"
                ],

                row[
                    "approved_documents"
                ],

                row[
                    "customer_pending"
                ],

                row[
                    "internal_pending"
                ],

                row[
                    "progress_percent"
                ],
            ]
        )


    write_report_sheet(
        workbook,
        "Sales Order Summary",
        sales_order_headers,
        sales_order_excel_rows,
    )


    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)


    filename = (
        "VDRL_Management_Report_"
        f"{timezone.localdate():%Y%m%d}"
        ".xlsx"
    )


    response = HttpResponse(
        output.getvalue(),

        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


    response[
        "Content-Disposition"
    ] = (
        f'attachment; filename="{filename}"'
    )

    record_audit_event(
    action=(
        AuditLog
        .Action
        .EXPORT
    ),

    actor=request.user,

    request=request,

    model_label=(
        "core.ManagementReport"
    ),

    object_id=(
        timezone.localdate()
        .isoformat()
    ),

    object_repr=(
        "VDRL Management Excel Report"
    ),

    description=(
        "Exported the filtered VDRL "
        "management report."
    ),

    event_data={
        "filters": (
            get_report_filters(
                request
            )
        ),

        "filename": filename,
    },
)

    return response